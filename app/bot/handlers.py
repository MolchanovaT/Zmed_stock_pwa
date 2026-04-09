import asyncio
import io
import logging
import os
import smtplib
import tempfile
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from math import ceil
from typing import Any

from aiogram import Router, F
from aiogram.filters import CommandStart, Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.types.input_file import BufferedInputFile
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy import select, func

from app.db.models import Stock, Cart, CartItem
from app.db.session import AsyncSessionLocal
from .states import Form, CartFlow
from zoneinfo import ZoneInfo

# Регистрируем шрифт
pdfmetrics.registerFont(TTFont("DejaVuSans", "app/fonts/DejaVuSans.ttf"))
pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "app/fonts/DejaVuSans-Bold.ttf"))

r = Router()

ROWS_PER_PAGE = 10

MD_SPECIAL = r"_*[]()~`>#+-=|{}.!"

FILTER_MAP = {
    "group": Stock.group_name,
    "region": Stock.region,
    "warehouse": Stock.warehouse,
    "category": Stock.category,
    "manufacturer": Stock.manufacturer,
    "brand": Stock.brand,
    "nom_type": Stock.nom_type,
}

ART_W = 15  # Артикул
NAME_W = 40  # Номенклатура
CHAR_W = 20  # Характеристика
BAL_W = 8  # Остаток

MSK = ZoneInfo("Europe/Moscow")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def esc(text: str) -> str:
    """Экранировать спецсимволы для Markdown V2."""
    return ''.join(f"\\{c}" if c in MD_SPECIAL else c for c in text)


def get_breadcrumb_text(data: dict, current_step: str) -> str:
    labels = {
        "group": "Склад",
        "region": "Регион",
        "warehouse": "Склад внутри региона",
        "category": "Категория",
        "manufacturer": "Производитель",
        "brand": "Марка",
        "nom_type": "Вид номенклатуры",
    }
    order = ["group", "region", "warehouse", "category", "manufacturer", "brand", "nom_type"]
    text = ""
    for field in order:
        val = data.get(field)
        if val is not None:
            text += f"{labels[field]}: {val}\n"
        if field == current_step:
            break
    return text.strip()


async def store_list(state: FSMContext, key: str, items: list[str]):
    full = ["все"] + sorted(set(items))
    await state.update_data(**{f"{key}_list": full})


async def build_group_keyboard(state: FSMContext, page: int = 1):
    """
    Получаем список групп складов из БД и формируем клавиатуру
    (в state он кэшируется, чтобы не ходить в БД каждый раз).
    """
    data = await state.get_data()
    if "group_list" not in data:  # ещё не кэшировали
        async with AsyncSessionLocal() as s:
            groups = await uniq("group_name", s)  # уже без None
        await store_list(state, "group", groups)
        data = await state.get_data()  # перечитать

    groups = data["group_list"]
    return paginated_keyboard(groups, page, "group")


def get_from_list(data: dict, key: str, index: int) -> Any | None:
    values = data.get(f"{key}_search_list") or data.get(f"{key}_list", [])
    if 0 <= index < len(values):
        return values[index]
    return None


def paginated_keyboard(
        items: list[str],
        page: int = 1,
        prefix: str = "",
        back_prefix: str = None,
        per_page: int = 10
) -> InlineKeyboardMarkup:
    total_pages = max(1, ceil(len(items) / per_page))
    page = max(1, min(page, total_pages))
    start_page = (page - 1) * per_page
    end = start_page + per_page
    current_items = items[start_page:end]

    rows = []

    for idx, item in enumerate(current_items):
        real_index = start_page + idx
        if item.lower() == "все":
            continue  # ⛔️ не добавляем в основной список
        rows.append([InlineKeyboardButton(text=item, callback_data=f"{prefix}_id:{real_index}")])

    # 🔄 Добавляем кнопку "все" вручную (один раз)
    if "все" in items:
        index_all = items.index("все")
        rows.insert(0, [InlineKeyboardButton(text="все", callback_data=f"{prefix}_id:{index_all}")])

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"{prefix}_id:page:{page - 1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"{prefix}_id:page:{page + 1}"))
    if nav:
        rows.append(nav)

    rows.append([InlineKeyboardButton(text="🔍 Поиск", callback_data=f"{prefix}_search")])

    if back_prefix:
        rows.append([InlineKeyboardButton(text="↩ Назад", callback_data=f"back:{back_prefix}")])

    rows.append([InlineKeyboardButton(text="🏠 В начало", callback_data="to_start")])

    return InlineKeyboardMarkup(inline_keyboard=rows)


@r.callback_query(F.data.endswith("_search"))
async def search_prompt_handler(c: CallbackQuery, state: FSMContext):
    prefix = c.data.split("_search")[0]
    await state.update_data(search_mode=prefix)
    await c.message.answer(f"Введите часть названия для поиска по: {prefix}")


def search_back_step(prefix: str) -> str:
    return {
        "group": "",
        "region": "group",
        "warehouse": "region",
        "category": "warehouse",
        "manufacturer": "category",
        "brand": "manufacturer",
        "nom_type": "brand"
    }.get(prefix, "")


@r.message(CommandStart())
async def cmd_start(msg: Message, state: FSMContext):
    await state.clear()
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🚀 Начать поиск", callback_data="begin")],
            [InlineKeyboardButton(text="🧺 Корзина", callback_data="open_cart")],
        ]
    )
    await msg.answer("Добро пожаловать!\nНажмите кнопку, чтобы начать поиск остатков:", reply_markup=kb)


@r.callback_query(F.data == "begin")
async def cb_begin(c: CallbackQuery, state: FSMContext):
    # то, что раньше делал /start
    kb = await build_group_keyboard(state, 1)
    await c.message.edit_text("1️⃣ Выберите группу складов:", reply_markup=kb)
    await state.set_state(Form.group)


# ─── «Новый поиск» из результатов ─────────────────────────────────────
@r.callback_query(F.data == "restart")
async def cb_restart(c: CallbackQuery, state: FSMContext):
    await cmd_start(c.message, state)


@r.callback_query(F.data == "to_start")
async def cb_to_start(c: CallbackQuery, state: FSMContext):
    await state.clear()
    kb = InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text="🚀 Начать поиск", callback_data="begin")]]
    )
    await c.message.answer("Добро пожаловать!\nНажмите кнопку, чтобы начать поиск остатков:", reply_markup=kb)


def result_nav_keyboard(current_page: int, total_pages: int) -> InlineKeyboardMarkup:
    nav = []
    if current_page > 1:
        nav.append(InlineKeyboardButton(text="◀️ Назад", callback_data=f"page:{current_page - 1}"))
    if current_page < total_pages:
        nav.append(InlineKeyboardButton(text="Вперёд ▶️", callback_data=f"page:{current_page + 1}"))

    row1 = nav or [InlineKeyboardButton(text="—", callback_data="noop")]
    row1 += [
        InlineKeyboardButton(text="🔍 Поиск", callback_data="search"),
        InlineKeyboardButton(text="📄 Скачать PDF", callback_data="download_pdf"),
    ]

    row2 = [
        InlineKeyboardButton(text="🛒 Оформить заказ", callback_data="order_checkout"),
        InlineKeyboardButton(text="🧺 Корзина", callback_data="open_cart"),
    ]

    row3 = [InlineKeyboardButton(text="🔄 Новый поиск", callback_data="restart")]

    return InlineKeyboardMarkup(inline_keyboard=[row1, row2, row3])


async def uniq(col: str, session, **f):
    stmt = select(getattr(Stock, col)).distinct()
    for k, v in f.items():
        if v and v.lower() != "все":
            stmt = stmt.filter(getattr(Stock, k) == v)

    res = await session.scalars(stmt)
    # ✨ убираем None и уже потом сортируем
    values = [x for x in res.all() if x is not None]
    values = [v for v in values if str(v).strip().lower() != "итого"]
    return sorted(values)


# @r.message(CommandStart())
# async def start(msg: Message, state: FSMContext):
#     await state.clear()
#     kb = await build_group_keyboard(state, 1)
#     await msg.answer("1️⃣ Выберите группу складов:", reply_markup=kb)
#     await state.set_state(Form.group)


@r.callback_query(Form.group, F.data.startswith("group_id:"))
async def handle_group(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("group_id:")[1]
    data = await state.get_data()

    # постраничная навигация
    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        kb = await build_group_keyboard(state, page)
        await c.message.edit_text("1️⃣ Выберите группу складов:", reply_markup=kb)
        return

    idx = int(payload)
    group = get_from_list(data, "group", idx)
    if group is None:
        await c.answer("Ошибка группы")
        return

    await state.update_data(group=group)

    # дальше всё как было (строим список регионов и переходим на Form.region)
    async with AsyncSessionLocal() as s:
        regions = await uniq("region", s, group_name=None if group == "все" else group)

    await store_list(state, "region", regions)
    data = await state.get_data()
    region_kb = paginated_keyboard(
        data["region_list"], 1, "region", "group"
    )

    await c.message.edit_text(
        f"Склад: {group}\n\n2️⃣ Выберите регион:",
        reply_markup=region_kb
    )
    await state.set_state(Form.region)


@r.callback_query(Form.region, F.data.startswith("region_id:"))
async def handle_region(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("region_id:")[1]
    data = await state.get_data()

    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        region_list = data.get("region_list", [])
        await c.message.edit_text(
            f"{get_breadcrumb_text(data, 'group')}\n\n2️⃣ Выберите регион:",
            reply_markup=paginated_keyboard(region_list, page, "region", "group"))
        return

    idx = int(payload)
    region = get_from_list(data, "region", idx)
    if region is None:
        await c.answer("Ошибка: регион не найден")
        return

    await state.update_data(region=region)

    async with AsyncSessionLocal() as s:
        warehouses = await uniq("warehouse", s, group_name=data["group"],
                                region=None if region == "все" else region)

    await store_list(state, "warehouse", warehouses)
    data = await state.get_data()
    warehouse_list = data.get("warehouse_list", [])

    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'region')}\n\n3️⃣ Выберите склад:",
        reply_markup=paginated_keyboard(warehouse_list, 1, "warehouse", "region")
    )
    await state.set_state(Form.warehouse)


@r.callback_query(Form.warehouse, F.data.startswith("warehouse_id:"))
async def handle_warehouse(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("warehouse_id:")[1]
    data = await state.get_data()

    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        warehouse_list = data.get("warehouse_list", [])
        await c.message.edit_text(
            f"{get_breadcrumb_text(data, 'region')}\n\n3️⃣ Выберите склад:",
            reply_markup=paginated_keyboard(warehouse_list, page, "warehouse", "region"))
        return

    idx = int(payload)
    warehouse = get_from_list(data, "warehouse", idx)
    if warehouse is None:
        await c.answer("Ошибка склада")
        return

    await state.update_data(warehouse=warehouse)

    async with AsyncSessionLocal() as s:
        cats = await uniq("category", s,
                          group_name=data["group"],
                          region=None if data["region"] == "все" else data["region"],
                          warehouse=None if warehouse == "все" else warehouse)

    await store_list(state, "category", cats)
    data = await state.get_data()
    category_list = data.get("category_list", [])

    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'warehouse')}\n\n4️⃣ Выберите вид:",
        reply_markup=paginated_keyboard(category_list, 1, "category", "warehouse")
    )
    await state.set_state(Form.category)


@r.callback_query(Form.category, F.data.startswith("category_id:"))
async def handle_category(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("category_id:")[1]
    data = await state.get_data()

    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        category_list = data.get("category_list", [])
        await c.message.edit_text(
            f"{get_breadcrumb_text(data, 'warehouse')}\n\n4️⃣ Выберите вид:",
            reply_markup=paginated_keyboard(category_list, page, "category", "warehouse"))
        return

    idx = int(payload)
    category = get_from_list(data, "category", idx)
    if category is None:
        await c.answer("Ошибка вида")
        return

    await state.update_data(category=category)

    async with AsyncSessionLocal() as s:
        mans = await uniq("manufacturer", s,
                          group_name=data["group"],
                          region=None if data["region"] == "все" else data["region"],
                          warehouse=None if data["warehouse"] == "все" else data["warehouse"],
                          category=None if category == "все" else category)

    await store_list(state, "manufacturer", mans)
    data = await state.get_data()
    manufacturer_list = data.get("manufacturer_list", [])

    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'category')}\n\n5️⃣ Выберите производителя:",
        reply_markup=paginated_keyboard(manufacturer_list, 1, "manufacturer", "category")
    )
    await state.set_state(Form.manufacturer)


@r.callback_query(Form.manufacturer, F.data.startswith("manufacturer_id:"))
async def handle_manufacturer(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("manufacturer_id:")[1]
    data = await state.get_data()

    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        manufacturer_list = data.get("manufacturer_list", [])
        await c.message.edit_text(
            f"{get_breadcrumb_text(data, 'category')}\n\n5️⃣ Выберите производителя:",
            reply_markup=paginated_keyboard(manufacturer_list, page, "manufacturer", "category"))
        return

    idx = int(payload)
    manufacturer = get_from_list(data, "manufacturer", idx)
    if manufacturer is None:
        await c.answer("Ошибка производителя")
        return

    await state.update_data(manufacturer=manufacturer)

    async with AsyncSessionLocal() as s:
        brands = await uniq("brand", s,
                            group_name=data["group"],
                            region=None if data["region"] == "все" else data["region"],
                            warehouse=None if data["warehouse"] == "все" else data["warehouse"],
                            category=None if data["category"] == "все" else data["category"],
                            manufacturer=None if manufacturer == "все" else manufacturer)

    await store_list(state, "brand", brands)
    data = await state.get_data()
    brand_list = data.get("brand_list", [])

    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'manufacturer')}\n\n6️⃣ Выберите марку (бренд):",
        reply_markup=paginated_keyboard(brand_list, 1, "brand", "manufacturer")
    )
    await state.set_state(Form.brand)


@r.callback_query(Form.brand, F.data.startswith("brand_id:"))
async def handle_brand(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("brand_id:")[1]
    data = await state.get_data()

    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        brand_list = data.get("brand_list", [])
        await c.message.edit_text(
            f"{get_breadcrumb_text(data, 'manufacturer')}\n\n6️⃣ Выберите марку (бренд):",
            reply_markup=paginated_keyboard(brand_list, page, "brand", "manufacturer"))
        return

    idx = int(payload)
    brand = get_from_list(data, "brand", idx)
    if brand is None:
        await c.answer("Ошибка бренда")
        return

    await state.update_data(brand=brand, page=1, search=None)

    async with AsyncSessionLocal() as s:
        nts = await uniq("nom_type", s,
                         group_name=data["group"],
                         region=None if data["region"] == "все" else data["region"],
                         warehouse=None if data["warehouse"] == "все" else data["warehouse"],
                         category=None if data["category"] == "все" else data["category"],
                         manufacturer=None if data["manufacturer"] == "все" else data["manufacturer"],
                         brand=None if brand == "все" else brand)

    await store_list(state, "nom_type", nts)
    nt_list = (await state.get_data())["nom_type_list"]

    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'manufacturer')}\n\n7️⃣ Выберите вид номенклатуры:",
        reply_markup=paginated_keyboard(nt_list, 1, "nom_type", "brand")
    )
    await state.set_state(Form.nom_type)


@r.callback_query(Form.nom_type, F.data.startswith("nom_type_id:"))
async def handle_nom_type(c: CallbackQuery, state: FSMContext):
    payload = c.data.split("nom_type_id:")[1]
    data = await state.get_data()

    if payload.startswith("page:"):
        page = int(payload.split(":")[1])
        nt_list = data.get("nom_type_list", [])
        await c.message.edit_text(
            f"{get_breadcrumb_text(data, 'manufacturer')}\n\n7️⃣ Выберите вид номенклатуры:",
            reply_markup=paginated_keyboard(nt_list, page, "nom_type", "brand"))
        return

    idx = int(payload)
    nom_type = get_from_list(data, "nom_type", idx)
    if nom_type is None:
        await c.answer("Ошибка вида номенклатуры")
        return

    await state.update_data(nom_type=nom_type, page=1, search=None)
    await state.set_state(Form.result_page)
    await show_result(c, state)


@r.callback_query(F.data.startswith("back:"))
async def go_back(c: CallbackQuery, state: FSMContext):
    step = c.data.split("back:")[1]
    data = await state.get_data()

    # Очистим возможный *_search_list
    await state.update_data({
        f"{step}_search_list": None,
        "search_mode": None
    })

    async with AsyncSessionLocal():
        match step:
            case "group":
                await cmd_start(c.message, state)

            case "region":
                regions = data.get("region_list", [])
                await c.message.edit_text(
                    f"{get_breadcrumb_text(data, 'group')}\n\n2️⃣ Выберите регион:",
                    reply_markup=paginated_keyboard(regions, 1, "region", "group"))
                await state.set_state(Form.region)

            case "warehouse":
                warehouses = data.get("warehouse_list", [])
                await c.message.edit_text(
                    f"{get_breadcrumb_text(data, 'region')}\n\n3️⃣ Выберите склад:",
                    reply_markup=paginated_keyboard(warehouses, 1, "warehouse", "region"))
                await state.set_state(Form.warehouse)

            case "category":
                cats = data.get("category_list", [])
                await c.message.edit_text(
                    f"{get_breadcrumb_text(data, 'warehouse')}\n\n4️⃣ Выберите вид:",
                    reply_markup=paginated_keyboard(cats, 1, "category", "warehouse"))
                await state.set_state(Form.category)

            case "manufacturer":
                mans = data.get("manufacturer_list", [])
                await c.message.edit_text(
                    f"{get_breadcrumb_text(data, 'category')}\n\n5️⃣ Выберите производителя:",
                    reply_markup=paginated_keyboard(mans, 1, "manufacturer", "category"))
                await state.set_state(Form.manufacturer)

            case "brand":
                brands = data.get("brand_list", [])
                await c.message.edit_text(
                    f"{get_breadcrumb_text(data, 'manufacturer')}\n\n6️⃣ Выберите марку (бренд):",
                    reply_markup=paginated_keyboard(brands, 1, "brand", "manufacturer"))
                await state.set_state(Form.brand)

            case "nom_type":
                nts = data.get("nom_type_list", [])
                await c.message.edit_text(
                    f"{get_breadcrumb_text(data, 'manufacturer')}\n\n7️⃣ Выберите вид номенклатуры:",
                    reply_markup=paginated_keyboard(nts, 1, "nom_type", "brand"))
                await state.set_state(Form.nom_type)


@r.callback_query(Form.group, F.data.startswith("group_page:"))
async def handle_group_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    kb = await build_group_keyboard(state, page)
    await c.message.edit_text("1️⃣ Выберите группу складов:", reply_markup=kb)


@r.callback_query(Form.region, F.data.startswith("region_page:"))
async def handle_region_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    regions = data.get("region_list", [])
    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'group')}\n\n2️⃣ Выберите регион:",
        reply_markup=paginated_keyboard(regions, page, "region", "group")
    )


@r.callback_query(Form.warehouse, F.data.startswith("warehouse_page:"))
async def handle_warehouse_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    warehouses = data.get("warehouse_list", [])
    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'region')}\n\n3️⃣ Выберите склад:",
        reply_markup=paginated_keyboard(warehouses, page, "warehouse", "region")
    )


@r.callback_query(Form.category, F.data.startswith("category_page:"))
async def handle_category_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    categories = data.get("category_list", [])
    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'warehouse')}\n\n4️⃣ Выберите вид:",
        reply_markup=paginated_keyboard(categories, page, "category", "warehouse")
    )


@r.callback_query(Form.manufacturer, F.data.startswith("manufacturer_page:"))
async def handle_manufacturer_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    manufacturers = data.get("manufacturer_list", [])
    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'category')}\n\n5️⃣ Выберите производителя:",
        reply_markup=paginated_keyboard(manufacturers, page, "manufacturer", "category")
    )


@r.callback_query(Form.brand, F.data.startswith("brand_page:"))
async def handle_brand_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    brands = data.get("brand_list", [])
    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'manufacturer')}\n\n6️⃣ Выберите марку (бренд):",
        reply_markup=paginated_keyboard(brands, page, "brand", "manufacturer")
    )


@r.callback_query(Form.nom_type, F.data.startswith("nom_type_page:"))
async def handle_nom_type_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    nt_list = data.get("nom_type_list", [])
    await c.message.edit_text(
        f"{get_breadcrumb_text(data, 'manufacturer')}\n\n7️⃣ Выберите вид номенклатуры:",
        reply_markup=paginated_keyboard(nt_list, page, "nom_type", "brand")
    )


# ─────────────────────────────────────────────────────────────
async def show_result(c: CallbackQuery | Message, state: FSMContext):
    """
    Выводит страницу остатков + дату-актуальности (берём из Stock.updated_at,
    куда при импорте заносится дата/время создания исходного файла).
    """
    data = await state.get_data()
    page = data.get("page", 1)
    search = data.get("search")

    async with AsyncSessionLocal() as s:
        # ── сами остатки
        stmt = (
            select(
                Stock.article,
                Stock.nomenclature,
                Stock.characteristic,
                func.sum(Stock.balance).label("bal"),

                # ➜ добавили: максимальный штамп внутри каждой группы
                func.max(Stock.updated_at).label("ts"),
            )
            .where(Stock.nomenclature.is_not(None))
            .group_by(
                Stock.article,
                Stock.nomenclature,
                Stock.characteristic,
            )
        )

        for key, column in FILTER_MAP.items():
            val = data.get(key)
            if val and val.lower() != "все":
                stmt = stmt.filter(column == val)

        if search:
            like = f"%{search}%"
            stmt = stmt.filter(
                Stock.nomenclature.ilike(like) |
                Stock.article.ilike(like) |
                Stock.characteristic.ilike(like)
            )

        all_rows = (await s.execute(stmt)).all()

        # ── максимальная updated_at (дата файла-источника)
        if all_rows:
            max_ts = max(row.ts for row in all_rows)  # aware-datetime (MSK уже не нужен)
            ts_str = max_ts.strftime("%d.%m.%Y %H:%M")
        else:
            ts_str = "–"

    # ── пагинация
    total = len(all_rows)
    total_pages = max(1, (total + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)
    page = max(1, min(page, total_pages))
    start_i = (page - 1) * ROWS_PER_PAGE
    end_i = start_i + ROWS_PER_PAGE
    chunk = all_rows[start_i:end_i]

    # ── текстовая таблица
    header = (
        f"{'Артикул':<{ART_W}} "
        f"{'Номенклатура':<{NAME_W}} "
        f"{'Характеристика':<{CHAR_W}} "
        f"{'Ост':>{BAL_W}}"
    )
    line = (
        f"{'-' * ART_W} "
        f"{'-' * NAME_W} "
        f"{'-' * CHAR_W} "
        f"{'-' * BAL_W}"
    )
    body = [
        f"{art[:ART_W]:<{ART_W}} "
        f"{n[:NAME_W]:<{NAME_W}} "
        f"{ch[:CHAR_W]:<{CHAR_W}} "
        f"{bal:>{BAL_W},.0f}"
        for art, n, ch, bal, _ in chunk
    ]

    breadcrumbs = esc(get_breadcrumb_text(data, "nom_type"))
    ts_str_md = esc(ts_str)

    text = (
            f"{breadcrumbs}\n"
            f"*Актуально на:* {ts_str_md}\n\n"
            f"```text\n"
            f"Страница {page} из {total_pages}\n"
            f"{header}\n{line}\n" +
            "\n".join(body) +
            "\n```"
    )

    markup = result_nav_keyboard(page, total_pages)

    if isinstance(c, CallbackQuery):
        await c.message.edit_text(text,
                                  parse_mode="MarkdownV2",
                                  reply_markup=markup)
    else:
        await c.answer(text,
                       parse_mode="MarkdownV2",
                       reply_markup=markup)


# ─────────────────────────────────────────────────────────────
# 2) PDF-отчёт
@r.callback_query(Form.result_page, F.data == "download_pdf")
async def download_pdf(c: CallbackQuery, state: FSMContext):
    data = await state.get_data()

    # какие колонки выводим
    select_fields = [Stock.region, Stock.warehouse,
                     Stock.article, Stock.nomenclature, Stock.characteristic,
                     func.sum(Stock.balance).label("bal")]
    group_by_cols = [Stock.region, Stock.warehouse,
                     Stock.article, Stock.nomenclature, Stock.characteristic]

    async with AsyncSessionLocal() as s:
        # --- данные об остатках
        stmt = select(*select_fields).group_by(*group_by_cols)

        for key, column in FILTER_MAP.items():
            val = data.get(key)
            if val and val.lower() != "все":
                stmt = stmt.filter(column == val)

        if data.get("search"):
            like = f"%{data['search']}%"
            stmt = stmt.filter(
                Stock.nomenclature.ilike(like) |
                Stock.article.ilike(like) |
                Stock.characteristic.ilike(like)
            )

        rows = (await s.execute(stmt)).all()

        # --- максимальная updated_at
        # отдельно берём «свежее» время точно так же,
        # как это сделано в show_result – надёжнее и без лишних select-полей
        ts_stmt = select(func.max(Stock.updated_at))
        for key, column in FILTER_MAP.items():
            val = data.get(key)
            if val and val.lower() != "все":
                ts_stmt = ts_stmt.filter(column == val)
        max_ts = await s.scalar(ts_stmt)
        ts_str = max_ts.strftime("%d.%m.%Y %H:%M") if max_ts else "–"

    # ─── PDF ───
    font_path = os.path.join("app", "fonts", "DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "app/fonts/DejaVuSans-Bold.ttf"))

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        doc = SimpleDocTemplate(tmp.name, pagesize=A4)
        styles = getSampleStyleSheet()

        normal = styles["Normal"]
        normal.fontName = "DejaVuSans"
        bold = ParagraphStyle(
            name="GroupHeader",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=11, leading=13)

        elems = []

        # ▸ шапка
        breadcrumbs = get_breadcrumb_text(data, "nom_type").replace("\n", "<br/>")
        elems.append(Paragraph(breadcrumbs, bold))
        elems.append(Paragraph(f"Актуально на: {ts_str}", bold))
        elems.append(Spacer(1, 8))

        # ▸ группировка Region / Warehouse
        from collections import defaultdict

        grouped = defaultdict(list)
        for row in rows:
            key = f"{row.region or '—'} / {row.warehouse or '—'}"
            title = f"{row.article or '—'}, {row.nomenclature or '—'}, {row.characteristic or '—'}"
            grouped[key].append((title, row.bal))

    for idx, (title, grp) in enumerate(grouped.items()):
        if idx:  # небольшой промежуток между группами
            elems.append(Spacer(1, 10))

        elems.append(Paragraph(title, bold))
        table_data = (
                [[Paragraph("Артикул, номенклатура, характеристика", normal),
                  Paragraph("Ост", normal)]] +
                [[Paragraph(title, normal), f"{b:,.0f}"] for title, b in grp]
        )

        t = Table(table_data, colWidths=[370, 80])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elems.append(t)

    doc.build(elems)

    with open(tmp.name, "rb") as f:
        await c.message.answer_document(
            BufferedInputFile(f.read(), filename="report.pdf"),
            caption="📄 PDF-отчёт с датой актуальности"
        )


@r.callback_query(Form.result_page, F.data.startswith("page:"))
async def change_page(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    await state.update_data(page=page)
    await show_result(c, state)


@r.callback_query(Form.result_page, F.data == "search")
async def ask_search(c: CallbackQuery, state: FSMContext):
    await c.message.answer("Введите часть названия для поиска по Номенклатуре:")
    await state.set_state(Form.result_page)


@r.message(Form.result_page)
async def search_query(msg: Message, state: FSMContext):
    await state.update_data(search=msg.text, page=1)
    await show_result(msg, state)


@r.message(
    F.text & ~F.text.startswith("/"),
    ~StateFilter(CartFlow.delivery_date, CartFlow.delivery_time, CartFlow.doctor),
)
async def handle_search_input(msg: Message, state: FSMContext):
    data = await state.get_data()

    prefix = data.get("search_mode")
    if not prefix:
        # Не в поиске — пропускаем сообщение, даём другим хендлерам (например, /start) обработать
        return

    original = data.get(f"{prefix}_list", [])
    query = msg.text.strip().lower()

    filtered = [item for item in original if query in item.lower()]
    if not filtered:
        await msg.answer("Ничего не найдено.")
        return

    await state.update_data({
        f"{prefix}_search_list": filtered,
        "search_mode": None
    })

    await msg.answer(
        f"Результаты поиска по: {query}",
        reply_markup=paginated_keyboard(filtered, 1, prefix, back_prefix=search_back_step(prefix))
    )


# ═══════════════════════════════════════════════════════════════════════════════
# КОРЗИНА И ОФОРМЛЕНИЕ ЗАКАЗА
# ═══════════════════════════════════════════════════════════════════════════════

ORDER_SELECT_PER_PAGE = 5
LPU_PER_PAGE = 10
CART_ITEMS_PER_PAGE = 5


# ── Вспомогательные функции ────────────────────────────────────────────────────

def build_order_card(
    items: list[dict],
    quantities: dict,   # str(idx) -> qty, qty=0 означает «не выбрано»
    idx: int,
) -> tuple[str, InlineKeyboardMarkup]:
    """Карточка одной позиции: полный текст + кнопки ➖/➕ для количества."""
    total = len(items)
    idx = max(0, min(idx, total - 1))
    item = items[idx]

    art  = item["article"] or "—"
    nom  = item["nomenclature"] or "—"
    char = item["characteristic"] or ""
    bal  = int(item["balance"])
    qty  = quantities.get(str(idx), 0)
    selected_count = sum(1 for v in quantities.values() if v > 0)

    lines = [
        f"Позиция {idx + 1} из {total}" +
        (f"  |  в заказе: {selected_count} поз." if selected_count else ""),
        "",
        f"Артикул:  {art}",
        nom,
    ]
    if char:
        lines.append(char)
    lines.append(f"Остаток: {bal} шт.")
    text = "\n".join(lines)

    nav = []
    if idx > 0:
        nav.append(InlineKeyboardButton(text="◀️ Пред", callback_data=f"ocard:{idx - 1}"))
    if idx < total - 1:
        nav.append(InlineKeyboardButton(text="➡️ След", callback_data=f"ocard:{idx + 1}"))

    rows = []
    if nav:
        rows.append(nav)
    rows.append([
        InlineKeyboardButton(text="➖", callback_data=f"ominus:{idx}"),
        InlineKeyboardButton(text=f"{qty} шт.", callback_data="noop"),
        InlineKeyboardButton(text="➕", callback_data=f"oplus:{idx}"),
    ])
    add_text = (
        f"✅ Добавить в корзину ({selected_count} поз.)"
        if selected_count else "✅ Добавить в корзину"
    )
    rows.append([InlineKeyboardButton(text=add_text, callback_data="order_add_selected")])
    rows.append([
        InlineKeyboardButton(text="🧺 Корзина", callback_data="open_cart"),
        InlineKeyboardButton(text="↩ К результатам", callback_data="back_to_results"),
    ])
    return text, InlineKeyboardMarkup(inline_keyboard=rows)


def lpu_keyboard(warehouses: list[str], page: int = 1) -> InlineKeyboardMarkup:
    """Клавиатура выбора ЛПУ из списка складов."""
    total = len(warehouses)
    total_pages = max(1, ceil(total / LPU_PER_PAGE))
    page = max(1, min(page, total_pages))
    start = (page - 1) * LPU_PER_PAGE
    end = min(start + LPU_PER_PAGE, total)

    rows = []
    for idx in range(start, end):
        w = warehouses[idx]
        rows.append([InlineKeyboardButton(text=w[:55], callback_data=f"lpu_sel:{idx}")])

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"lpu_page:{page - 1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"lpu_page:{page + 1}"))
    if nav:
        rows.append(nav)

    rows.append([InlineKeyboardButton(text="✏️ Ввести вручную", callback_data="lpu_manual")])
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="order_cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def build_cart_view(
    cart: Cart,
    items: list[CartItem],
    page: int = 1,
) -> tuple[str, InlineKeyboardMarkup]:
    """Текст и клавиатура просмотра корзины."""
    lpu_str = cart.lpu or "не указано"
    total = len(items)
    total_pages = max(1, ceil(total / CART_ITEMS_PER_PAGE))
    page = max(1, min(page, total_pages))
    start = (page - 1) * CART_ITEMS_PER_PAGE
    end = min(start + CART_ITEMS_PER_PAGE, total)

    if not items:
        text = f"🧺 Корзина пуста\nЛПУ: {lpu_str}"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔍 Продолжить поиск", callback_data="to_start")],
        ])
        return text, kb

    lines = [f"🧺 Корзина | ЛПУ: {lpu_str}", f"Заказ #{cart.id} | Позиций: {total}\n"]
    page_items = items[start:end]
    for i, item in enumerate(page_items, start + 1):
        art = (item.article or "")[:12]
        nom = (item.nomenclature or "")[:28]
        char = (item.characteristic or "")[:18]
        avail = int(item.available_balance or 0)
        lines.append(f"{i}. {art} | {nom}")
        lines.append(f"   {char}")
        lines.append(f"   Кол-во: {item.quantity}  |  Остаток на складе: {avail}")

    text = "\n".join(lines)

    rows = []
    for item in page_items:
        nom_short = (item.nomenclature or "")[:18]
        rows.append([
            InlineKeyboardButton(text=f"🗑 {nom_short}", callback_data=f"ci_del:{item.id}"),
        ])
        rows.append([
            InlineKeyboardButton(text="➖", callback_data=f"ci_minus:{item.id}"),
            InlineKeyboardButton(text=str(item.quantity), callback_data="noop"),
            InlineKeyboardButton(text="➕", callback_data=f"ci_plus:{item.id}"),
        ])

    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="◀️", callback_data=f"cart_page:{page - 1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="▶️", callback_data=f"cart_page:{page + 1}"))
    if nav:
        rows.append(nav)

    rows.append([InlineKeyboardButton(text="📤 Сформировать заказ", callback_data="cart_place_order")])
    rows.append([
        InlineKeyboardButton(text="🔍 Продолжить поиск", callback_data="to_start"),
        InlineKeyboardButton(text="🏠 В начало", callback_data="back_to_start"),
    ])
    return text, InlineKeyboardMarkup(inline_keyboard=rows)


def _send_email_sync(
    subject: str,
    cart_id: int,
    lpu: str,
    user_full_name: str,
    user_username: str,
    user_tg_id: int,
    now_str: str,
    items_snapshot: list[tuple],
    delivery_date: str = "не указано",
    delivery_time: str = "не указано",
    doctor: str = "не указано",
    instrument: str = "нет",
) -> None:
    """Синхронная отправка письма: HTML-таблица в теле + Excel во вложении."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.config import SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SMTP_FROM, ORDER_EMAIL_TO

    if not SMTP_HOST or not ORDER_EMAIL_TO:
        logger.info("SMTP не настроен, письмо не отправлено.")
        return

    recipients = [r.strip() for r in ORDER_EMAIL_TO.split(",") if r.strip()]

    # ── HTML-таблица ───────────────────────────────────────────────
    rows_html = ""
    for i, (art, nom, char, qty, _avail) in enumerate(items_snapshot, 1):
        bg = "#ffffff" if i % 2 == 0 else "#f4f8fd"
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{i}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd">{art}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd">{nom}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd">{char}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{qty}</td>'
            f'</tr>'
        )

    html_body = f"""<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#333">
<h2 style="color:#2c5f8a">Заказ #{cart_id}</h2>
<table style="border-collapse:collapse;margin-bottom:16px">
  <tr><td style="padding:4px 16px 4px 0;color:#666">Дата:</td><td><b>{now_str}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">ЛПУ:</td><td><b>{lpu}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Дата доставки:</td><td><b>{delivery_date}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Время доставки:</td><td><b>{delivery_time}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Врач:</td><td><b>{doctor}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Инструмент:</td><td><b>{instrument}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Пользователь:</td>
      <td><b>{user_full_name}</b> (@{user_username}, ID: {user_tg_id})</td></tr>
</table>
<table style="border-collapse:collapse;width:100%">
  <thead>
    <tr style="background:#2c5f8a;color:#fff">
      <th style="padding:8px 10px;border:1px solid #1e4a73">№</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Артикул</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Номенклатура</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Характеристика</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Кол-во</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>
</body></html>"""

    # ── Excel-вложение ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заказ {cart_id}"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2C5F8A")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="AAAAAA")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center")

    # Шапка документа — каждый реквизит на отдельной строке
    label_font = Font(bold=True)
    meta = [
        ("Заказ №",        str(cart_id)),
        ("Дата",           now_str),
        ("ЛПУ",            lpu),
        ("Дата доставки",  delivery_date),
        ("Время доставки", delivery_time),
        ("Врач",           doctor),
        ("Инструмент",     instrument),
        ("Пользователь",   f"{user_full_name} (@{user_username}, ID: {user_tg_id})"),
    ]
    for row_num, (label, value) in enumerate(meta, 1):
        ws.merge_cells(f"A{row_num}:E{row_num}")
        ws.cell(row=row_num, column=1, value=f"{label}:  {value}").font = label_font
        ws.row_dimensions[row_num].height = 18

    table_start = len(meta) + 2  # пустая строка-разделитель перед таблицей

    # Шапка таблицы
    for col, header in enumerate(["№", "Артикул", "Номенклатура", "Характеристика", "Кол-во"], 1):
        c = ws.cell(row=table_start, column=col, value=header)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, brd
    ws.row_dimensions[table_start].height = 20

    # Данные
    for i, (art, nom, char, qty, _avail) in enumerate(items_snapshot, 1):
        for col, val in enumerate([i, art, nom, char, qty], 1):
            c = ws.cell(row=table_start + i, column=col, value=val)
            c.border = brd
            if col in (1, 5):
                c.alignment = center

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    # ── Письмо ────────────────────────────────────────────────────
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SMTP_FROM or SMTP_USER
    msg["To"]      = ", ".join(recipients)

    alt = MIMEMultipart("alternative")
    plain = (
        f"Заказ #{cart_id}\nДата: {now_str}\nЛПУ: {lpu}\n"
        f"Дата доставки: {delivery_date}\nВремя доставки: {delivery_time}\n"
        f"Врач: {doctor}\nИнструмент: {instrument}\n"
        f"Пользователь: {user_full_name} (@{user_username}, ID: {user_tg_id})\n\n"
        + "\n".join(
            f"{i}. [{art}] {nom} | {char} — {qty} шт."
            for i, (art, nom, char, qty, _) in enumerate(items_snapshot, 1)
        )
    )
    alt.attach(MIMEText(plain, "plain", "utf-8"))
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    filename = f"Заказ_{cart_id}_{now_str.replace(':', '-').replace(' ', '_')}.xlsx"
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(msg["From"], recipients, msg.as_string())
        logger.info("Заказ отправлен на email: %s", recipients)
    except Exception as e:
        logger.error("Ошибка отправки email: %s", e)


async def send_order_notification(
    subject: str,
    cart_id: int,
    lpu: str,
    user_full_name: str,
    user_username: str,
    user_tg_id: int,
    now_str: str,
    items_snapshot: list[tuple],
    delivery_date: str = "не указано",
    delivery_time: str = "не указано",
    doctor: str = "не указано",
    instrument: str = "нет",
) -> None:
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None, _send_email_sync,
        subject, cart_id, lpu, user_full_name, user_username, user_tg_id, now_str, items_snapshot,
        delivery_date, delivery_time, doctor, instrument,
    )


async def _fetch_order_items_from_state(state: FSMContext):
    """Достаём все строки результатов из БД (повторяет логику show_result)."""
    data = await state.get_data()
    async with AsyncSessionLocal() as s:
        stmt = (
            select(
                Stock.article,
                Stock.nomenclature,
                Stock.characteristic,
                func.sum(Stock.balance).label("bal"),
            )
            .where(Stock.nomenclature.is_not(None))
            .group_by(Stock.article, Stock.nomenclature, Stock.characteristic)
        )
        for key, column in FILTER_MAP.items():
            val = data.get(key)
            if val and val.lower() != "все":
                stmt = stmt.filter(column == val)
        search = data.get("search")
        if search:
            like = f"%{search}%"
            stmt = stmt.filter(
                Stock.nomenclature.ilike(like) |
                Stock.article.ilike(like) |
                Stock.characteristic.ilike(like)
            )
        rows = (await s.execute(stmt)).all()
    return [
        {
            "article": row.article or "",
            "nomenclature": row.nomenclature or "",
            "characteristic": row.characteristic or "",
            "balance": float(row.bal or 0),
        }
        for row in rows
    ]


async def _do_add_items_to_cart(
    event: CallbackQuery | Message,
    state: FSMContext,
    lpu: str,
    cart_id: int | None,
) -> None:
    """Добавляет выбранные позиции в корзину (новую или существующую)."""
    data = await state.get_data()
    order_items: list[dict] = data.get("order_items", [])
    quantities: dict = data.get("order_quantities", {})
    selected = [int(k) for k, v in quantities.items() if v > 0]
    user_id = event.from_user.id

    async with AsyncSessionLocal() as s:
        if cart_id is None:
            cart = Cart(tg_user_id=user_id, lpu=lpu, status="active")
            s.add(cart)
            await s.flush()
            cart_id = cart.id

        for idx in selected:
            if 0 <= idx < len(order_items):
                item = order_items[idx]
                s.add(CartItem(
                    cart_id=cart_id,
                    article=item["article"],
                    nomenclature=item["nomenclature"],
                    characteristic=item["characteristic"],
                    quantity=quantities.get(str(idx), 1),
                    available_balance=item["balance"],
                ))
        await s.commit()

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🧺 Открыть корзину", callback_data="open_cart")],
        [InlineKeyboardButton(text="🔍 Продолжить поиск", callback_data="to_start")],
    ])
    text = f"✅ {len(selected)} поз. добавлено в корзину!\nЛПУ: {lpu}"
    if isinstance(event, CallbackQuery):
        await event.message.edit_text(text, reply_markup=kb)
    else:
        await event.answer(text, reply_markup=kb)
    await state.set_state(Form.result_page)


async def _show_lpu_selection(c: CallbackQuery, state: FSMContext) -> None:
    """Показывает список складов для выбора ЛПУ."""
    async with AsyncSessionLocal() as s:
        warehouses = await uniq("warehouse", s)
    await state.update_data(lpu_list=warehouses, lpu_page=1)
    await state.set_state(CartFlow.lpu_select)
    await c.message.edit_text(
        "Выберите ЛПУ (учреждение):",
        reply_markup=lpu_keyboard(warehouses, 1),
    )


async def _show_cart(
    event: CallbackQuery | Message,
    state: FSMContext,
    page: int = 1,
    edit: bool = False,
) -> None:
    """Загружает активную корзину пользователя и показывает её."""
    user_id = event.from_user.id
    async with AsyncSessionLocal() as s:
        result = await s.execute(
            select(Cart)
            .where(Cart.tg_user_id == user_id, Cart.status == "active")
            .order_by(Cart.created_at.desc())
        )
        cart = result.scalars().first()
        if not cart:
            no_cart_kb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🔍 Начать поиск", callback_data="begin")],
            ])
            if isinstance(event, CallbackQuery):
                await event.message.answer("🧺 Корзина пуста — начните поиск.", reply_markup=no_cart_kb)
            else:
                await event.answer("🧺 Корзина пуста — начните поиск.", reply_markup=no_cart_kb)
            return

        cart_id = cart.id
        cart_lpu = cart.lpu
        # Явный SELECT без кэша — берём свежие данные из БД
        items_res = await s.execute(
            select(CartItem)
            .where(CartItem.cart_id == cart_id)
            .order_by(CartItem.id)
            .execution_options(populate_existing=True)
        )
        items_data = [
            {
                "id": it.id,
                "article": it.article,
                "nomenclature": it.nomenclature,
                "characteristic": it.characteristic,
                "quantity": it.quantity,
                "available_balance": it.available_balance,
            }
            for it in items_res.scalars().all()
        ]
        logger.info("_show_cart: cart_id=%s items=%s", cart_id, [(d["id"], d["quantity"]) for d in items_data])

    # Создаём простые объекты-заглушки для build_cart_view
    class _Cart:
        def __init__(self, id_, lpu):
            self.id = id_
            self.lpu = lpu

    class _Item:
        def __init__(self, d):
            self.id = d["id"]
            self.article = d["article"]
            self.nomenclature = d["nomenclature"]
            self.characteristic = d["characteristic"]
            self.quantity = d["quantity"]
            self.available_balance = d["available_balance"]

    cart_obj = _Cart(cart_id, cart_lpu)
    items_obj = [_Item(d) for d in items_data]

    await state.update_data(active_cart_id=cart_id, cart_page=page)
    await state.set_state(CartFlow.cart_view)
    text, kb = build_cart_view(cart_obj, items_obj, page)

    if isinstance(event, CallbackQuery):
        if edit:
            await event.message.edit_text(text, reply_markup=kb)
        else:
            await event.message.answer(text, reply_markup=kb)
    else:
        await event.answer(text, reply_markup=kb)


# ── Хэндлеры ──────────────────────────────────────────────────────────────────

@r.callback_query(Form.result_page, F.data == "order_checkout")
async def order_checkout(c: CallbackQuery, state: FSMContext):
    """Открывает экран выбора позиций для заказа."""
    order_items = await _fetch_order_items_from_state(state)
    if not order_items:
        await c.answer("Нет позиций для заказа.")
        return
    await state.update_data(order_items=order_items, order_quantities={}, order_current=0)
    await state.set_state(CartFlow.select_items)
    text, kb = build_order_card(order_items, {}, 0)
    await c.message.edit_text(text, reply_markup=kb)


@r.callback_query(CartFlow.select_items, F.data.startswith("oplus:"))
async def order_qty_plus(c: CallbackQuery, state: FSMContext):
    idx = int(c.data.split(":")[1])
    data = await state.get_data()
    quantities: dict = dict(data.get("order_quantities", {}))
    order_items = data.get("order_items", [])
    bal = int(order_items[idx]["balance"]) if idx < len(order_items) else 9999
    current = quantities.get(str(idx), 0)
    quantities[str(idx)] = min(current + 1, bal)
    await state.update_data(order_quantities=quantities)
    text, kb = build_order_card(order_items, quantities, idx)
    await c.message.edit_text(text, reply_markup=kb)
    await c.answer()


@r.callback_query(CartFlow.select_items, F.data.startswith("ominus:"))
async def order_qty_minus(c: CallbackQuery, state: FSMContext):
    idx = int(c.data.split(":")[1])
    data = await state.get_data()
    quantities: dict = dict(data.get("order_quantities", {}))
    order_items = data.get("order_items", [])
    current = quantities.get(str(idx), 0)
    if current > 0:
        quantities[str(idx)] = current - 1
    await state.update_data(order_quantities=quantities)
    text, kb = build_order_card(order_items, quantities, idx)
    await c.message.edit_text(text, reply_markup=kb)
    await c.answer()


@r.callback_query(CartFlow.select_items, F.data.startswith("ocard:"))
async def order_navigate_card(c: CallbackQuery, state: FSMContext):
    """Переход к соседней позиции."""
    idx = int(c.data.split(":")[1])
    data = await state.get_data()
    await state.update_data(order_current=idx)
    order_items = data.get("order_items", [])
    quantities = data.get("order_quantities", {})
    text, kb = build_order_card(order_items, quantities, idx)
    await c.message.edit_text(text, reply_markup=kb)
    await c.answer()


@r.callback_query(CartFlow.select_items, F.data == "order_add_selected")
async def order_add_selected(c: CallbackQuery, state: FSMContext):
    """Проверяет наличие корзины и спрашивает: добавить к существующей или новую."""
    data = await state.get_data()
    quantities: dict = data.get("order_quantities", {})
    selected = [int(k) for k, v in quantities.items() if v > 0]
    if not selected:
        await c.answer("Укажите количество хотя бы для одной позиции!", show_alert=True)
        return

    user_id = c.from_user.id
    async with AsyncSessionLocal() as s:
        result = await s.execute(
            select(Cart)
            .where(Cart.tg_user_id == user_id, Cart.status == "active")
            .order_by(Cart.created_at.desc())
        )
        existing = result.scalars().first()

    if existing:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(
                text=f"➕ Добавить к заказу #{existing.id} (ЛПУ: {existing.lpu or '—'})",
                callback_data=f"cart_add_to:{existing.id}",
            )],
            [InlineKeyboardButton(text="🆕 Создать новый заказ", callback_data="cart_new")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="order_cancel")],
        ])
        await c.message.edit_text(
            f"У вас есть активный заказ #{existing.id} (ЛПУ: {existing.lpu or 'не указано'}).\n\n"
            "Добавить позиции к нему или создать новый заказ?",
            reply_markup=kb,
        )
    else:
        await _show_lpu_selection(c, state)


@r.callback_query(CartFlow.select_items, F.data.startswith("cart_add_to:"))
async def cart_add_to_existing(c: CallbackQuery, state: FSMContext):
    """Добавляет позиции к уже существующей корзине."""
    cart_id = int(c.data.split(":")[1])
    async with AsyncSessionLocal() as s:
        cart = await s.get(Cart, cart_id)
        if not cart:
            await c.answer("Заказ не найден.")
            return
        cart_lpu = cart.lpu or ""
    await _do_add_items_to_cart(c, state, cart_lpu, cart_id)


@r.callback_query(CartFlow.select_items, F.data == "cart_new")
async def cart_new_handler(c: CallbackQuery, state: FSMContext):
    """Создаёт новый заказ — запускает выбор ЛПУ."""
    await _show_lpu_selection(c, state)


@r.callback_query(CartFlow.lpu_select, F.data.startswith("lpu_sel:"))
async def lpu_selected(c: CallbackQuery, state: FSMContext):
    idx = int(c.data.split(":")[1])
    data = await state.get_data()
    lpu_list = data.get("lpu_list", [])
    if not (0 <= idx < len(lpu_list)):
        await c.answer("Ошибка выбора ЛПУ.")
        return
    lpu = lpu_list[idx]
    await _do_add_items_to_cart(c, state, lpu, cart_id=None)


@r.callback_query(CartFlow.lpu_select, F.data.startswith("lpu_page:"))
async def lpu_page_handler(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    data = await state.get_data()
    lpu_list = data.get("lpu_list", [])
    await state.update_data(lpu_page=page)
    await c.message.edit_reply_markup(reply_markup=lpu_keyboard(lpu_list, page))
    await c.answer()


@r.callback_query(CartFlow.lpu_select, F.data == "lpu_manual")
async def lpu_manual_input(c: CallbackQuery, state: FSMContext):
    await c.message.edit_text("Введите название ЛПУ:")
    await state.set_state(CartFlow.lpu_input)


@r.message(CartFlow.lpu_input)
async def lpu_text_input(msg: Message, state: FSMContext):
    lpu = msg.text.strip()
    await _do_add_items_to_cart(msg, state, lpu, cart_id=None)


# ── Корзина ────────────────────────────────────────────────────────────────────

@r.callback_query(F.data == "open_cart")
async def open_cart(c: CallbackQuery, state: FSMContext):
    await _show_cart(c, state, page=1, edit=False)


@r.message(Command("cart"))
async def cmd_cart(msg: Message, state: FSMContext):
    await _show_cart(msg, state, page=1, edit=False)


@r.callback_query(F.data.startswith("cart_page:"))
async def cart_page_handler(c: CallbackQuery, state: FSMContext):
    page = int(c.data.split(":")[1])
    await _show_cart(c, state, page=page, edit=True)


@r.callback_query(F.data.startswith("ci_del:"))
async def cart_item_delete(c: CallbackQuery, state: FSMContext):
    item_id = int(c.data.split(":")[1])
    data = await state.get_data()
    cart_page = data.get("cart_page", 1)
    async with AsyncSessionLocal() as s:
        item = await s.get(CartItem, item_id)
        if item:
            await s.delete(item)
            await s.commit()
    await c.answer("Позиция удалена.")
    await _show_cart(c, state, page=cart_page, edit=True)


@r.callback_query(F.data.startswith("ci_minus:"))
async def cart_item_minus(c: CallbackQuery, state: FSMContext):
    item_id = int(c.data.split(":")[1])
    data = await state.get_data()
    cart_page = data.get("cart_page", 1)
    async with AsyncSessionLocal() as s:
        item = await s.get(CartItem, item_id)
        if item and item.quantity > 1:
            item.quantity -= 1
            s.add(item)
            await s.commit()
    await c.answer()
    await _show_cart(c, state, page=cart_page, edit=True)


@r.callback_query(F.data.startswith("ci_plus:"))
async def cart_item_plus(c: CallbackQuery, state: FSMContext):
    item_id = int(c.data.split(":")[1])
    data = await state.get_data()
    cart_page = data.get("cart_page", 1)
    async with AsyncSessionLocal() as s:
        item = await s.get(CartItem, item_id)
        if item:
            item.quantity += 1
            s.add(item)
            await s.commit()
    await c.answer()
    await _show_cart(c, state, page=cart_page, edit=True)


@r.callback_query(F.data == "noop")
async def noop_handler(c: CallbackQuery):
    await c.answer()


@r.callback_query(F.data == "cart_place_order")
async def cart_place_order(c: CallbackQuery, state: FSMContext):
    """Начинает оформление заказа: проверяет корзину и запрашивает дату доставки."""
    data = await state.get_data()
    cart_id = data.get("active_cart_id")

    async with AsyncSessionLocal() as s:
        cart = await s.get(Cart, cart_id)
        if not cart:
            await c.answer("Заказ не найден.")
            return
        items_res = await s.execute(
            select(CartItem).where(CartItem.cart_id == cart_id).order_by(CartItem.id)
        )
        items = list(items_res.scalars().all())
        if not items:
            await c.answer("Корзина пуста!", show_alert=True)
            return

    await state.set_state(CartFlow.delivery_date)
    await c.message.edit_text("📅 Укажите дату доставки (например: 28.03.2026):")


@r.message(CartFlow.delivery_date)
async def delivery_date_input(msg: Message, state: FSMContext):
    text = msg.text.strip()
    try:
        delivery_dt = datetime.strptime(text, "%d.%m.%Y").date()
    except ValueError:
        await msg.answer("❌ Неверный формат даты. Введите в формате ДД.ММ.ГГГГ, например: 28.03.2026")
        return
    if delivery_dt < datetime.now().date():
        await msg.answer("❌ Дата доставки не может быть в прошлом. Введите корректную дату:")
        return
    await state.update_data(delivery_date=text)
    await state.set_state(CartFlow.delivery_time)
    await msg.answer("⏰ Укажите время доставки (например: 10:00):")


@r.message(CartFlow.delivery_time)
async def delivery_time_input(msg: Message, state: FSMContext):
    text = msg.text.strip()
    try:
        datetime.strptime(text, "%H:%M")
    except ValueError:
        await msg.answer("❌ Неверный формат времени. Введите в формате ЧЧ:ММ, например: 10:00")
        return
    await state.update_data(delivery_time=text)
    await state.set_state(CartFlow.doctor)
    await msg.answer("👨‍⚕️ Укажите врача (контактное лицо):")


@r.message(CartFlow.doctor)
async def doctor_input(msg: Message, state: FSMContext):
    await state.update_data(doctor=msg.text.strip())
    await state.set_state(CartFlow.instrument)
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Да", callback_data="instrument:да"),
        InlineKeyboardButton(text="❌ Нет", callback_data="instrument:нет"),
    ]])
    await msg.answer("🔧 Нужен инструмент?", reply_markup=kb)


@r.callback_query(CartFlow.instrument, F.data.startswith("instrument:"))
async def instrument_selected(c: CallbackQuery, state: FSMContext):
    instrument = c.data.split(":")[1]
    await state.update_data(instrument=instrument)
    await c.answer()
    await _submit_order(c, state)


async def _submit_order(c: CallbackQuery, state: FSMContext) -> None:
    """Финальный шаг: сохраняет заказ и отправляет уведомление."""
    data = await state.get_data()
    cart_id = data.get("active_cart_id")
    delivery_date = data.get("delivery_date", "не указано")
    delivery_time = data.get("delivery_time", "не указано")
    doctor = data.get("doctor", "не указано")
    instrument = data.get("instrument", "нет")

    async with AsyncSessionLocal() as s:
        cart = await s.get(Cart, cart_id)
        if not cart:
            await c.message.answer("Заказ не найден.")
            return
        items_res = await s.execute(
            select(CartItem).where(CartItem.cart_id == cart_id).order_by(CartItem.id)
        )
        items = list(items_res.scalars().all())
        if not items:
            await c.message.answer("Корзина пуста!")
            return
        cart_id_val = cart.id
        cart_lpu_val = cart.lpu or "не указано"
        items_snapshot = [
            (it.article, it.nomenclature, it.characteristic, it.quantity, int(it.available_balance or 0))
            for it in items
        ]
        cart.status = "submitted"
        cart.delivery_date = delivery_date
        cart.delivery_time = delivery_time
        cart.doctor = doctor
        cart.instrument = instrument
        await s.commit()

    user = c.from_user
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    order_lines = [
        f"Заказ #{cart_id_val}",
        f"Дата: {now_str}",
        f"Пользователь: {user.full_name} (@{user.username or 'N/A'}, ID: {user.id})",
        f"ЛПУ: {cart_lpu_val}",
        f"Дата доставки: {delivery_date}",
        f"Время доставки: {delivery_time}",
        f"Врач: {doctor}",
        f"Инструмент: {instrument}",
        "",
        "Позиции:",
    ]
    for i, (art, nom, char, qty, avail) in enumerate(items_snapshot, 1):
        order_lines.append(
            f"{i}. [{art}] {nom} | {char} — {qty} шт. (остаток: {avail})"
        )

    order_text = "\n".join(order_lines)
    subject = f"Заказ #{cart_id_val} от {user.full_name} | ЛПУ: {cart_lpu_val} | {now_str}"

    asyncio.create_task(send_order_notification(
        subject,
        cart_id_val,
        cart_lpu_val,
        user.full_name,
        user.username or "N/A",
        user.id,
        now_str,
        items_snapshot,
        delivery_date,
        delivery_time,
        doctor,
        instrument,
    ))

    # TODO: интеграция с 1С (добавить вызов API 1С здесь)
    logger.info("Заказ #%d оформлен:\n%s", cart_id_val, order_text)

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔍 Новый поиск", callback_data="to_start")],
    ])
    await c.message.answer(
        f"✅ Заказ #{cart_id_val} оформлен!\n\n{order_text}\n\n"
        "Уведомление отправлено.",
        reply_markup=kb,
    )
    await state.clear()


# ── Вспомогательные колбэки ───────────────────────────────────────────────────

@r.callback_query(F.data == "back_to_results")
async def back_to_results(c: CallbackQuery, state: FSMContext):
    await state.set_state(Form.result_page)
    await show_result(c, state)


@r.callback_query(F.data == "order_cancel")
async def order_cancel(c: CallbackQuery, state: FSMContext):
    await state.set_state(Form.result_page)
    await show_result(c, state)


@r.callback_query(F.data == "back_to_start")
async def back_to_start(c: CallbackQuery, state: FSMContext):
    await state.clear()
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Начать поиск", callback_data="begin")],
        [InlineKeyboardButton(text="🧺 Корзина", callback_data="open_cart")],
    ])
    await c.message.answer("Добро пожаловать!\nНажмите кнопку, чтобы начать поиск остатков:", reply_markup=kb)
