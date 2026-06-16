"""Email-уведомления о заказах из PWA-корзины."""

import asyncio
import io
import logging
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

logger = logging.getLogger(__name__)


def _send_email_sync(
    subject: str,
    cart_id: int,
    lpu: str,
    user_full_name: str,
    user_username: str,
    user_tg_id: int,
    now_str: str,
    items_snapshot: list[tuple],
    delivery_date: str = "не указано",
    delivery_time: str = "не указано",
    doctor: str = "не указано",
    instrument: str = "нет",
    source_lpu: str = "не указано",
    comment: str = "",
    kind: str = "implants",
    user_email: str = "",
) -> None:
    """Синхронная отправка письма: HTML-таблица в теле + Excel во вложении."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy.orm import Session
    from app.config import SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SMTP_FROM
    from app.db.session import sync_engine
    from app.db.models import EmailRecipient

    if not SMTP_HOST:
        logger.info("SMTP не настроен, письмо не отправлено.")
        return

    # Список адресов рассылки берём из БД (редактируется в админке).
    with Session(sync_engine) as s:
        recipients = [r.email.strip() for r in s.query(EmailRecipient).all() if r.email and r.email.strip()]

    # Копия письма заказчику, если у него указан email и его ещё нет в списке.
    ue = (user_email or "").strip()
    if ue and "@" in ue and ue.lower() not in {r.lower() for r in recipients}:
        recipients.append(ue)

    if not recipients:
        logger.info("Нет адресатов рассылки, письмо не отправлено.")
        return

    kind_label = "инструменты" if kind == "supplies" else "импланты"

    rows_html = ""
    for i, (art, nom, char, qty, _avail) in enumerate(items_snapshot, 1):
        bg = "#ffffff" if i % 2 == 0 else "#f4f8fd"
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{i}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd">{art}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd">{nom}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd">{char}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{qty}</td>'
            f'</tr>'
        )

    comment_html_row = (
        f'<tr><td style="padding:4px 16px 4px 0;color:#666">Комментарий:</td>'
        f'<td><b>{comment}</b></td></tr>'
        if comment else ""
    )
    html_body = f"""<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#333">
<h2 style="color:#2c5f8a">Заказ #{cart_id} ({kind_label})</h2>
<table style="border-collapse:collapse;margin-bottom:16px">
  <tr><td style="padding:4px 16px 4px 0;color:#666">Дата:</td><td><b>{now_str}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Склад отбора:</td><td><b>{source_lpu}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">ЛПУ-получатель:</td><td><b>{lpu}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Дата доставки:</td><td><b>{delivery_date}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Время доставки:</td><td><b>{delivery_time}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Врач:</td><td><b>{doctor}</b></td></tr>
  <tr><td style="padding:4px 16px 4px 0;color:#666">Инструмент:</td><td><b>{instrument}</b></td></tr>
  {comment_html_row}
  <tr><td style="padding:4px 16px 4px 0;color:#666">Пользователь:</td>
      <td><b>{user_full_name}</b> (@{user_username}, ID: {user_tg_id})</td></tr>
</table>
<table style="border-collapse:collapse;width:100%">
  <thead>
    <tr style="background:#2c5f8a;color:#fff">
      <th style="padding:8px 10px;border:1px solid #1e4a73">№</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Артикул</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Номенклатура</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Характеристика</th>
      <th style="padding:8px 10px;border:1px solid #1e4a73">Кол-во</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>
</body></html>"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заказ {cart_id}"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2C5F8A")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="AAAAAA")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center")

    label_font = Font(bold=True)
    meta = [
        ("Заказ №",        f"{cart_id} ({kind_label})"),
        ("Дата",           now_str),
        ("Склад отбора",   source_lpu),
        ("ЛПУ-получатель", lpu),
        ("Дата доставки",  delivery_date),
        ("Время доставки", delivery_time),
        ("Врач",           doctor),
        ("Инструмент",     instrument),
    ]
    if comment:
        meta.append(("Комментарий", comment))
    meta.append(("Пользователь", f"{user_full_name} (@{user_username}, ID: {user_tg_id})"))
    for row_num, (label, value) in enumerate(meta, 1):
        ws.merge_cells(f"A{row_num}:E{row_num}")
        ws.cell(row=row_num, column=1, value=f"{label}:  {value}").font = label_font
        ws.row_dimensions[row_num].height = 18

    table_start = len(meta) + 2

    for col, header in enumerate(["№", "Артикул", "Номенклатура", "Характеристика", "Кол-во"], 1):
        c = ws.cell(row=table_start, column=col, value=header)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, brd
    ws.row_dimensions[table_start].height = 20

    for i, (art, nom, char, qty, _avail) in enumerate(items_snapshot, 1):
        for col, val in enumerate([i, art, nom, char, qty], 1):
            c = ws.cell(row=table_start + i, column=col, value=val)
            c.border = brd
            if col in (1, 5):
                c.alignment = center

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SMTP_FROM or SMTP_USER
    msg["To"]      = ", ".join(recipients)

    alt = MIMEMultipart("alternative")
    plain_comment = f"Комментарий: {comment}\n" if comment else ""
    plain = (
        f"Заказ #{cart_id} ({kind_label})\nДата: {now_str}\n"
        f"Склад отбора: {source_lpu}\nЛПУ-получатель: {lpu}\n"
        f"Дата доставки: {delivery_date}\nВремя доставки: {delivery_time}\n"
        f"Врач: {doctor}\nИнструмент: {instrument}\n"
        f"{plain_comment}"
        f"Пользователь: {user_full_name} (@{user_username}, ID: {user_tg_id})\n\n"
        + "\n".join(
            f"{i}. [{art}] {nom} | {char} — {qty} шт."
            for i, (art, nom, char, qty, _) in enumerate(items_snapshot, 1)
        )
    )
    alt.attach(MIMEText(plain, "plain", "utf-8"))
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    filename = f"Заказ_{cart_id}_{now_str.replace(':', '-').replace(' ', '_')}.xlsx"
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(msg["From"], recipients, msg.as_string())
        logger.info("Заказ отправлен на email: %s", recipients)
    except Exception as e:
        logger.error("Ошибка отправки email: %s", e)


async def send_order_notification(
    subject: str,
    cart_id: int,
    lpu: str,
    user_full_name: str,
    user_username: str,
    user_tg_id: int,
    now_str: str,
    items_snapshot: list[tuple],
    delivery_date: str = "не указано",
    delivery_time: str = "не указано",
    doctor: str = "не указано",
    instrument: str = "нет",
    source_lpu: str = "не указано",
    comment: str = "",
    kind: str = "implants",
    user_email: str = "",
) -> None:
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None, _send_email_sync,
        subject, cart_id, lpu, user_full_name, user_username, user_tg_id, now_str, items_snapshot,
        delivery_date, delivery_time, doctor, instrument, source_lpu, comment, kind, user_email,
    )
